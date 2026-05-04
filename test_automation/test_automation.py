from playwright.sync_api import sync_playwright
import time
import os
import argparse
import re
from pathlib import Path
import sys
import json
import base64
import openpyxl
from openpyxl.cell.cell import MergedCell
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.backends import default_backend
from cryptography.hazmat.primitives import padding as aes_padding

# Configuration
ROOT_DIR = Path(__file__).resolve().parent.parent
TESTS_DIR = ROOT_DIR / "test_automation"
DATA_DIR  = ROOT_DIR / "data"

DEFAULT_EXCEL_CANDIDATES = [
    str(DATA_DIR  / "Assignment 1 - Test cases.xlsx"),
    str(TESTS_DIR / "Assignment 1 - Test cases.xlsx"),
]

DEFAULT_SHEET_NAME = " Test cases"
DEFAULT_FRONTEND_URL = os.getenv("FRONTEND_URL", "https://www.pixelssuite.com/chat-translator")

DEFAULT_INPUT_COLUMN_CANDIDATES = [
    "Singlish",
    "Input",
    "Singlish Input",
    "Test Input",
    "Source",
    "Sentence",
    "Text",
]

DEFAULT_EXPECTED_COLUMN_CANDIDATES = [
    "Sinhala",
    "Expected_Output",
    "Expected Output",
    "Expected output",
    "Expected",
    "Expected Sinhala",
]

DEFAULT_ACTUAL_COLUMN_CANDIDATES = [
    "Actual_Output",
    "Actual Output",
    "Actual output",
    "Actual",
]

DEFAULT_STATUS_COLUMN_CANDIDATES = [
    "Status",
    "Result",
    "Pass/Fail",
    "Pass Fail",
]

# ---------------------------------------------------------------------------
# AES-256-CBC helpers  (same key/IV as the site's CryptoJS bundle)
# Key  : "12345678901234567890123456789012"  (32 bytes UTF-8)
# IV   : "1234567890123456"                   (16 bytes UTF-8)
# ---------------------------------------------------------------------------
_AES_KEY  = b'12345678901234567890123456789012'
_AES_IV   = b'1234567890123456'
_API_HOST   = 'sw5794hs568nw7-8000.proxy.runpod.net'
_CHAT_PATH  = '/chat/translate'       # plain JSON  – used by Chat Translator UI
_STD_PATH   = '/transliterate'        # AES-256-CBC – used by Standard Sinhala UI


def _aes_decrypt(base64_str: str):
    """Decrypt AES-256-CBC base64 string; return parsed JSON dict (or None)."""
    try:
        ct = base64.b64decode(base64_str)
        cipher = Cipher(algorithms.AES(_AES_KEY), modes.CBC(_AES_IV), backend=default_backend())
        dec = cipher.decryptor()
        padded = dec.update(ct) + dec.finalize()
        unpadder = aes_padding.PKCS7(128).unpadder()
        plain = unpadder.update(padded) + unpadder.finalize()
        return json.loads(plain.decode('utf-8'))
    except Exception as exc:
        print(f'  [AES] Decrypt error: {exc}')
        return None


def _extract_sinhala(decrypted) -> str:
    """Pull Sinhala text from the decrypted API response object."""
    if not decrypted:
        return ''
    if isinstance(decrypted, str):
        return decrypted
    if decrypted.get('output'):
        return str(decrypted['output'])
    segs = decrypted.get('segments', [])
    if segs:
        return ''.join(s.get('text', '') for s in segs if s)
    return ''


DEFAULT_WAIT_MS = 5000
DEFAULT_RETRIES = 8
DEFAULT_RETRY_WAIT_MS = 1000
DEFAULT_TYPE_DELAY_MS = 30
DEFAULT_TIMEOUT_MS = 60000
DEFAULT_SLOW_MO_MS = 0

def _configure_stdout():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

def _pick_existing_path(candidates):
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return candidates[0] if candidates else None

def _resolve_path(p: str | None) -> str | None:
    if not p:
        return None
    path = Path(p)
    if path.is_absolute():
        return str(path)
    root_candidate = (ROOT_DIR / path).resolve()
    if root_candidate.exists():
        return str(root_candidate)
    tests_candidate = (TESTS_DIR / path).resolve()
    if tests_candidate.exists():
        return str(tests_candidate)
    return str(root_candidate)

def _normalize_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def _header_values(ws, row_index: int) -> list:
    max_col = max(1, int(ws.max_column or 1))
    return [ws.cell(row=row_index, column=c).value for c in range(1, max_col + 1)]

def _find_header_row(ws, max_scan_rows: int) -> int:
    input_tokens = {_normalize_header(v) for v in DEFAULT_INPUT_COLUMN_CANDIDATES}
    expected_tokens = {_normalize_header(v) for v in DEFAULT_EXPECTED_COLUMN_CANDIDATES}
    actual_tokens = {_normalize_header(v) for v in DEFAULT_ACTUAL_COLUMN_CANDIDATES}
    status_tokens = {_normalize_header(v) for v in DEFAULT_STATUS_COLUMN_CANDIDATES}

    best_score = -1
    best_row = 1
    scan_limit = max(1, min(int(max_scan_rows), int(ws.max_row or 1)))
    for r in range(1, scan_limit + 1):
        values = _header_values(ws, r)
        texts = [v for v in values if isinstance(v, str) and v.strip() and len(v.strip()) <= 40]
        if len(texts) < 2:
            continue

        norms = {_normalize_header(v) for v in texts}
        if "tcid" in norms and "input" in norms and "expectedoutput" in norms:
            return r

        if "input" not in norms:
            continue
        if not (norms & expected_tokens):
            continue

        score = 0
        for v in texts:
            n = _normalize_header(v)
            if n in input_tokens:
                score += 3
            if n in expected_tokens:
                score += 2
            if n in actual_tokens:
                score += 1
            if n in status_tokens:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    return best_row

def _merged_top_left_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return cell
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(row=rng.min_row, column=rng.min_col)
    return ws.cell(row=row, column=col)

def _is_top_left_of_merged_cell(ws, row: int, col: int) -> bool:
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        return True
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng.min_row == row and rng.min_col == col
    return True

def _set_cell_value(ws, row: int, col: int, value):
    cell = _merged_top_left_cell(ws, row, col)
    cell.value = value

def _find_column_index(header_values: list, requested_name: str | None, candidates: list[str]) -> int | None:
    indexed = []
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        indexed.append((i, str(v)))

    norm_to_index: dict[str, int] = {}
    for i, v in indexed:
        n = _normalize_header(v)
        if n and n not in norm_to_index:
            norm_to_index[n] = i

    def match(name: str) -> int | None:
        n = _normalize_header(name)
        if not n:
            return None
        if n in norm_to_index:
            return norm_to_index[n]
        for i, v in indexed:
            if n in _normalize_header(v) or _normalize_header(v) in n:
                return i
        return None

    if requested_name:
        found = match(requested_name)
        if found:
            return found

    for c in candidates:
        found = match(c)
        if found:
            return found

    return None

def _last_header_col(header_values: list) -> int:
    last = 0
    for i, v in enumerate(header_values, start=1):
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        last = i
    return last

def _ensure_column(ws, header_row: int, header_values: list, desired_name: str) -> int:
    found = _find_column_index(header_values, desired_name, [])
    if found:
        return found
    col = _last_header_col(header_values) + 1
    ws.cell(row=header_row, column=col).value = desired_name
    if col <= len(header_values):
        header_values[col - 1] = desired_name
    else:
        while len(header_values) < col - 1:
            header_values.append(None)
        header_values.append(desired_name)
    return col

def _dismiss_overlays(page):
    candidates = [
        ("button", re.compile(r"^(Accept|I Agree|Agree|OK|Got it)$", re.IGNORECASE)),
        ("button", re.compile(r"^(Accept all|Accept All)$", re.IGNORECASE)),
    ]
    for role, name in candidates:
        try:
            btn = page.get_by_role(role, name=name).first
            if btn.is_visible():
                btn.click(timeout=2000)
                page.wait_for_timeout(500)
        except Exception:
            pass

def _clear_textarea(page, locator, attempts: int = 3):
    for _ in range(max(1, int(attempts))):
        try:
            locator.click(timeout=2000)
        except Exception:
            pass
        try:
            page.keyboard.press("Control+A")
            page.keyboard.press("Backspace")
        except Exception:
            pass
        try:
            locator.fill("")
        except Exception:
            pass
        try:
            if locator.input_value() == "":
                return
        except Exception:
            pass
        try:
            locator.evaluate(
                """(el) => { el.value = ''; el.dispatchEvent(new Event('input', { bubbles: true })); }"""
            )
            if locator.input_value() == "":
                return
        except Exception:
            pass
        page.wait_for_timeout(200)

def _ensure_input_value(page, input_locator, text: str, type_delay_ms: int):
    _clear_textarea(page, input_locator)
    if type_delay_ms and int(type_delay_ms) > 0:
        input_locator.click(timeout=2000)
        input_locator.type(text, delay=int(type_delay_ms))
    else:
        input_locator.fill(text)
    try:
        current = input_locator.input_value()
        if current is None:
            return
        if str(current).strip() == text.strip():
            return
    except Exception:
        return
    page.wait_for_timeout(150)
    _clear_textarea(page, input_locator)
    input_locator.fill(text)

def _read_output(is_chat: bool, output_locator) -> str:
    if is_chat:
        try:
            v = output_locator.input_value()
            if v is not None:
                v = str(v).strip()
                if v:
                    return v
        except Exception:
            pass
    try:
        v = output_locator.inner_text()
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    try:
        v = output_locator.text_content()
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    try:
        v = output_locator.evaluate("(el) => el && ('value' in el ? el.value : '')")
        if v is not None:
            v = str(v).strip()
            if v:
                return v
    except Exception:
        pass
    return ""

def _find_chat_locators(page, timeout_ms: int):
    deadline = time.time() + (max(1, timeout_ms) / 1000)
    last_debug = None
    while time.time() < deadline:
        _dismiss_overlays(page)
        try:
            input_by_ph = page.locator('textarea[placeholder*="English"]').first
            output_by_ph = page.locator('textarea[placeholder*="Sinhala"]').first
            if input_by_ph.count() > 0 and output_by_ph.count() > 0 and input_by_ph.is_visible() and output_by_ph.is_visible():
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return input_by_ph, output_by_ph, action
        except Exception:
            pass

        try:
            count = page.locator("textarea").count()
            visible = []
            for i in range(count):
                loc = page.locator("textarea").nth(i)
                if loc.is_visible():
                    visible.append(loc)
            if len(visible) >= 2:
                action = page.get_by_role("button", name=re.compile(r"^Transliterate$", re.IGNORECASE)).first
                return visible[0], visible[1], action
        except Exception as e:
            last_debug = str(e)

        page.wait_for_timeout(500)

    try:
        meta = page.evaluate(
            """() => Array.from(document.querySelectorAll('textarea')).map(t => ({
              placeholder: t.getAttribute('placeholder') || '',
              disabled: !!t.disabled,
              readOnly: !!t.readOnly,
              visible: !!(t.offsetParent)
            }))"""
        )
        print("Debug: textarea meta:", meta)
    except Exception as e:
        print("Debug: failed to read textarea meta:", e)
    if last_debug:
        print("Debug: last error:", last_debug)
    raise RuntimeError("Could not find Chat UI locators (input/output textareas).")

def _parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", default=_pick_existing_path(DEFAULT_EXCEL_CANDIDATES))
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=0)
    parser.add_argument("--max-header-scan-rows", type=int, default=30)
    parser.add_argument("--input-col", default=None)
    parser.add_argument("--expected-col", default=None)
    parser.add_argument("--actual-col", default=None)
    parser.add_argument("--status-col", default=None)
    parser.add_argument("--url", default=DEFAULT_FRONTEND_URL)
    parser.add_argument("--output", default=None)
    parser.add_argument("--save-every", type=int, default=0)
    parser.add_argument("--headless", action="store_true", default=False)
    parser.add_argument("--wait-ms", type=int, default=DEFAULT_WAIT_MS)
    parser.add_argument("--retries", type=int, default=DEFAULT_RETRIES)
    parser.add_argument("--retry-wait-ms", type=int, default=DEFAULT_RETRY_WAIT_MS)
    parser.add_argument("--type-delay-ms", type=int, default=DEFAULT_TYPE_DELAY_MS)
    parser.add_argument("--timeout-ms", type=int, default=DEFAULT_TIMEOUT_MS)
    parser.add_argument("--slow-mo-ms", type=int, default=DEFAULT_SLOW_MO_MS)
    parser.add_argument("--keep-open", action="store_true", default=False)
    return parser.parse_args()

def run_test():
    _configure_stdout()
    args = _parse_args()
    args.excel = _resolve_path(args.excel)
    if args.output:
        args.output = _resolve_path(args.output)
    else:
        # Write results back into the same source Excel file
        args.output = args.excel

    if not args.excel or not os.path.exists(args.excel):
        print(f"Error: File '{args.excel}' not found.")
        return

    try:
        wb = openpyxl.load_workbook(args.excel)
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return

    if args.sheet and args.sheet in wb.sheetnames:
        ws = wb[args.sheet]
    else:
        ws = wb.active

    header_row = int(args.header_row or 0)
    if header_row <= 0:
        header_row = _find_header_row(ws, int(args.max_header_scan_rows))

    header_values = _header_values(ws, header_row)

    input_col_idx = _find_column_index(header_values, args.input_col, DEFAULT_INPUT_COLUMN_CANDIDATES)
    expected_col_idx = _find_column_index(header_values, args.expected_col, DEFAULT_EXPECTED_COLUMN_CANDIDATES)

    if not input_col_idx:
        printable = [str(v) if v is not None else "" for v in header_values]
        print("Error: Could not resolve input column.")
        print(f"Header row: {header_row}")
        print(f"Available columns: {printable}")
        return

    actual_col_name = args.actual_col or "Actual output"
    status_col_name = args.status_col or "Status"

    actual_col_idx = _find_column_index(header_values, args.actual_col, DEFAULT_ACTUAL_COLUMN_CANDIDATES)
    status_col_idx = _find_column_index(header_values, args.status_col, DEFAULT_STATUS_COLUMN_CANDIDATES)

    actual_col_idx = actual_col_idx or _ensure_column(ws, header_row, header_values, actual_col_name)
    status_col_idx = status_col_idx or _ensure_column(ws, header_row, header_values, status_col_name)

    rows_total = max(0, int(ws.max_row or 0) - header_row)
    print(f"Starting Frontend-Only test with {rows_total} rows...")

    with sync_playwright() as p:
        # 2. Launch Browser
        if args.headless:
            print("Running in headless mode: browser UI will not be visible. Remove --headless to watch typing.")
        browser = p.chromium.launch(
            headless=args.headless,
            slow_mo=max(0, int(args.slow_mo_ms)),
            args=["--disable-blink-features=AutomationControlled"],
        )
        page = browser.new_page()
        # Hide navigator.webdriver to avoid bot detection
        page.add_init_script(
            """Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"""
        )
        page.set_default_timeout(max(1000, int(args.timeout_ms)))

        is_chat = "chat-translator" in (args.url or "")

        # API response predicate – matches both endpoints
        def _is_api_response(resp):
            return _API_HOST in resp.url and (_CHAT_PATH in resp.url or _STD_PATH in resp.url)

        # 4. Iterate Rows  (navigate fresh for each row to avoid stuck-button state)
        processed = 0
        for row_index in range(header_row + 1, int(ws.max_row or 0) + 1):
            if not _is_top_left_of_merged_cell(ws, row_index, input_col_idx):
                continue

            input_cell = _merged_top_left_cell(ws, row_index, input_col_idx)
            input_value = input_cell.value
            singlish_input = str(input_value).strip() if input_value is not None else ""
            if not singlish_input:
                continue

            expected_value = (
                _merged_top_left_cell(ws, row_index, expected_col_idx).value if expected_col_idx else None
            )
            expected_sinhala = str(expected_value).strip() if expected_value is not None else ""

            print(f"Testing [Row {row_index}]: {singlish_input}")

            try:
                # Navigate fresh each row — avoids the button getting stuck in
                # "Transliterating..." state from a previous test case.
                page.goto(args.url, wait_until="domcontentloaded")

                # Re-locate elements (fresh page, so locators are new)
                if is_chat:
                    try:
                        input_locator, output_locator, action_locator = _find_chat_locators(page, int(args.timeout_ms))
                    except Exception as loc_err:
                        print(f"  Could not find chat UI: {loc_err}")
                        _set_cell_value(ws, row_index, status_col_idx, 'UI Error')
                        if args.save_every and int(args.save_every) > 0:
                            wb.save(args.output)
                        continue
                else:
                    input_locator = page.locator("textarea")
                    output_locator = page.locator("div.card").filter(has_text=re.compile(r"\\bSinhala\\b")).locator("div.bg-slate-50").first
                    action_locator = None

                _dismiss_overlays(page)

                # Use expect_response to capture the API call triggered by typing/clicking.
                # The context manager must wrap ALL actions that could trigger the API call
                # (both fill and button click) to avoid missing early responses.
                actual_output = ""
                try:
                    with page.expect_response(_is_api_response, timeout=int(args.wait_ms)) as resp_info:
                        _ensure_input_value(page, input_locator, singlish_input, int(args.type_delay_ms))
                        if action_locator:
                            action_locator.click()

                    api_resp = resp_info.value
                    print(f'  [API] Response: HTTP {api_resp.status} {api_resp.url}')
                    body_json = api_resp.json()

                    if _CHAT_PATH in api_resp.url:
                        # /chat/translate — plain JSON {"input":..., "output":...}
                        actual_output = str(body_json.get('output', ''))
                    elif _STD_PATH in api_resp.url and body_json.get('data'):
                        # /transliterate — AES-256-CBC encrypted
                        decrypted = _aes_decrypt(body_json['data'])
                        actual_output = _extract_sinhala(decrypted)

                    if actual_output:
                        print(f'  [API] Output: "{actual_output}"')
                    else:
                        print(f'  [API] Empty output. Body: {str(body_json)[:120]}')

                except Exception as api_err:
                    print(f'  [API] No response captured ({api_err}); reading textarea...')
                    if action_locator:
                        try:
                            action_locator.click()
                        except Exception:
                            pass
                    # Poll for output rather than waiting the full timeout upfront
                    deadline_fb = time.time() + max(0, int(args.wait_ms)) / 1000
                    while time.time() < deadline_fb:
                        cur = _read_output(is_chat, output_locator)
                        if cur:
                            actual_output = cur
                            break
                        page.wait_for_timeout(500)

                _set_cell_value(ws, row_index, actual_col_idx, actual_output)

                if expected_sinhala:
                    status = "PASS" if actual_output == expected_sinhala else "FAIL"
                else:
                    status = "COLLECTED"
                _set_cell_value(ws, row_index, status_col_idx, status)
                print(f"  -> {status}")
                processed += 1
                if args.save_every and int(args.save_every) > 0 and processed % int(args.save_every) == 0:
                    wb.save(args.output)
                
            except Exception as e:
                print(f"Error in UI interaction: {e}")
                try:
                    _set_cell_value(ws, row_index, status_col_idx, "UI Error")
                except Exception:
                    pass
                if args.save_every and int(args.save_every) > 0:
                    try:
                        wb.save(args.output)
                    except Exception:
                        pass

        if args.keep_open and not args.headless:
            try:
                wb.save(args.output)
            except Exception:
                pass
            print("Keeping browser open. Press CTRL+C to stop.")
            try:
                while True:
                    page.wait_for_timeout(1000)
            except KeyboardInterrupt:
                try:
                    wb.save(args.output)
                except Exception:
                    pass
        browser.close()

    try:
        wb.save(args.output)
    except Exception as e:
        print(f"Error saving output file '{args.output}': {e}")
        return

    print(f"Test completed. Results saved to {args.output}")

if __name__ == "__main__":
    run_test()
